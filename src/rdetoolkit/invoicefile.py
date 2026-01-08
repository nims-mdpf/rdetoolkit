from __future__ import annotations

import copy
import json
import os
import re
import shutil
import sys
import warnings
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Literal, Protocol, Union

from rdetoolkit import __version__
from rdetoolkit.exceptions import InvoiceSchemaValidationError, StructuredError
from rdetoolkit.fileops import readf_json, writef_json

if TYPE_CHECKING:
    import pandas as pd

    from rdetoolkit.models.invoice import (
        FixedHeaders,
        GeneralAttributeConfig,
        SpecificAttributeConfig,
        TemplateConfig,
    )
    from rdetoolkit.models.invoice_schema import SampleField
    from rdetoolkit.models.rde2types import RdeDatasetPaths, RdeFsPath, RdeOutputResourcePath

STATIC_DIR = Path(__file__).parent / "static"
EX_GENERALTERM = STATIC_DIR / "ex_generalterm.csv"
EX_SPECIFICTERM = STATIC_DIR / "ex_specificterm.csv"
MAGIC_VARIABLE_PATTERN = re.compile(r"\$\{([^{}]+)\}")


def _ensure_pandas() -> Any:
    import pandas as pd

    return pd


def _ensure_openpyxl_styles() -> tuple[Any, Any, Any]:
    from openpyxl.styles import Border, Font, Side

    return Border, Font, Side


def _ensure_openpyxl_utils() -> Any:
    from openpyxl.utils import get_column_letter

    return get_column_letter


def _ensure_chardet() -> Any:
    import chardet

    return chardet


def _ensure_validation_error() -> type[Exception]:
    from pydantic import ValidationError

    return ValidationError


def _ensure_logger() -> Callable[..., Any]:
    from rdetoolkit.rdelogger import get_logger

    return get_logger


logger = _ensure_logger()(__name__)


def read_excelinvoice(excelinvoice_filepath: RdeFsPath) -> tuple[pd.DataFrame, pd.DataFrame | None, pd.DataFrame | None]:
    """Deprecated wrapper around :class:`ExcelInvoiceFile`.

    This helper will be removed in version 1.5.0. Please instantiate ``ExcelInvoiceFile`` directly and use the
    ``dfexcelinvoice``, ``df_general``, and ``df_specific`` attributes.
    """
    warnings.warn(
        "read_excelinvoice() is deprecated and will be removed in version 1.5.0. "
        "Instantiate ExcelInvoiceFile(invoice_path) and use the .dfexcelinvoice, .df_general, and .df_specific attributes instead.",
        DeprecationWarning,
        stacklevel=2,
    )
    excel_invoice = ExcelInvoiceFile(Path(excelinvoice_filepath))
    return excel_invoice.dfexcelinvoice, excel_invoice.df_general, excel_invoice.df_specific


def _process_invoice_sheet(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    hd1 = list(df.iloc[1, :].fillna(""))
    hd2 = list(df.iloc[2, :].fillna(""))
    df.columns = [f"{s1}/{s2}" if s1 else s2 for s1, s2 in zip(hd1, hd2)]
    return df.iloc[4:, :].reset_index(drop=True).copy()


def _process_general_term_sheet(df: pd.DataFrame) -> pd.DataFrame:
    _df_general = df[1:].copy()
    _df_general.columns = ["term_id", "key_name"]
    return _df_general


def _process_specific_term_sheet(df: pd.DataFrame) -> pd.DataFrame:
    _df_specific = df[1:].copy()
    _df_specific.columns = ["sample_class_id", "term_id", "key_name"]
    return _df_specific


SheetType = Literal["invoice", "general_term", "specific_term", "unknown"]


def _identify_sheet_type(sh_name: str, df: pd.DataFrame) -> SheetType:
    """Identify the type of an Excel sheet.

    Args:
        sh_name: Excel sheet name.
        df: DataFrame loaded from the sheet.

    Returns:
        Sheet type:
        - "invoice": InvoiceList sheet (cell A1 is "invoiceList_format_id")
        - "general_term": generalTerm sheet
        - "specific_term": specificTerm sheet
        - "unknown": Other sheets (skipped in processing)
    """
    if not df.empty and len(df) > 0 and len(df.columns) > 0:
        target_comment_value = df.iat[0, 0]
        if target_comment_value == "invoiceList_format_id":
            return "invoice"

    if sh_name == "generalTerm":
        return "general_term"
    if sh_name == "specificTerm":
        return "specific_term"

    return "unknown"


# Sheet type to processor function mapping
_SHEET_PROCESSORS: dict[SheetType, Callable[[pd.DataFrame], pd.DataFrame]] = {
    "invoice": _process_invoice_sheet,
    "general_term": _process_general_term_sheet,
    "specific_term": _process_specific_term_sheet,
}


def check_exist_rawfiles(dfexcelinvoice: pd.DataFrame, excel_rawfiles: list[Path]) -> list[Path]:
    """Checks for the existence of raw file paths listed in a DataFrame against a list of file Paths.

    This function compares a set of file names extracted from the `data_file_names/name` column of the provided DataFrame (dfexcelinvoice) with the names of files in the excel_rawfiles list.
    If there are file names in the DataFrame that are not present in the excel_rawfiles list, it raises a StructuredError with a message indicating the missing file.
    If all file names in the DataFrame are present in the excel_rawfiles list, it returns a list of Path objects from excel_rawfiles, sorted in the order they appear in the DataFrame.

    Args:
        dfexcelinvoice (pd.DataFrame): A DataFrame containing file names in the 'data_file_names/name' column.
        excel_rawfiles (list[Path]): A list of Path objects representing file paths.

    Raises:
        tructuredError: If any file name in dfexcelinvoice is not found in excel_rawfiles.

    Returns:
        list[Path]: A list of Path objects corresponding to the file names in dfexcelinvoice, ordered as they appear in the DataFrame.
    """
    file_set_group = {f.name for f in excel_rawfiles}
    file_set_invoice = set(dfexcelinvoice["data_file_names/name"])
    if file_set_invoice - file_set_group:
        emsg = f"ERROR: raw file not found: {(file_set_invoice-file_set_group).pop()}"
        raise StructuredError(emsg)
    # Sort excel_rawfiles in the order they appear in the invoice
    _tmp = {f.name: f for f in excel_rawfiles}
    try:
        return [_tmp[f] for f in dfexcelinvoice["data_file_names/name"]]
    except KeyError as e:
        emsg = f"Invalid or missing key in data_file_names/name: {e}"
        raise StructuredError(emsg) from e


def _assign_invoice_val(invoiceobj: dict[str, Any], key1: str, key2: str, valobj: Any, invoiceschema_obj: dict[str, Any]) -> None:
    """When the destination key, which is the first key 'keys1', is 'custom', valobj is cast according to the invoiceschema_obj. In all other cases, valobj is assigned without changing its type."""
    if key1 == "custom":
        from rdetoolkit import rde2util

        dct_schema = invoiceschema_obj["properties"][key1]["properties"][key2]
        try:
            invoiceobj[key1][key2] = rde2util.castval(valobj, dct_schema["type"], dct_schema.get("format"))
        except StructuredError as struct_err:
            emsg = f"ERROR: failed to cast invoice value for key [{key1}][{key2}]"
            raise StructuredError(emsg) from struct_err
    else:
        invoiceobj[key1][key2] = valobj


def overwrite_invoicefile_for_dpfterm(
    invoiceobj: dict[str, Any],
    invoice_dst_filepath: RdeFsPath,
    invoiceschema_filepath: RdeFsPath,
    invoice_info: dict[str, Any],
) -> None:
    """A function to overwrite DPF metadata into an invoice file.

    Args:
        invoiceobj (dict[str, Any]): The object of invoice.json.
        invoice_dst_filepath (RdeFsPath): The file path for the destination invoice.json.
        invoiceschema_filepath (RdeFsPath): The file path of invoice.schema.json.
        invoice_info (dict[str, Any]): Information about the invoice file.
    """
    chardet = _ensure_chardet()
    with open(invoiceschema_filepath, "rb") as f:
        data = f.read()
    enc = chardet.detect(data)["encoding"]
    with open(invoiceschema_filepath, encoding=enc) as f:
        invoiceschema_obj = json.load(f)
    for k, v in invoice_info.items():
        _assign_invoice_val(invoiceobj, "custom", k, v, invoiceschema_obj)
    with open(invoice_dst_filepath, "w", encoding=enc) as fout:
        json.dump(invoiceobj, fout, indent=4, ensure_ascii=False)


def check_exist_rawfiles_for_folder(dfexcelinvoice: pd.DataFrame, rawfiles_tpl: tuple) -> list:
    """Function to check the existence of rawfiles_tpl specified for a folder.

    It checks whether rawfiles_tpl, specified as an index, exists in all indexes of ExcelInvoice.
    Assumes that the names of the terminal folders are unique and checks for the existence of rawfiles_tpl.

    Args:
        dfexcelinvoice (DataFrame): The dataframe of ExcelInvoice.
        rawfiles_tpl (tuple): Tuple of raw files.

    Returns:
        list: A list of rawfiles_tpl sorted in the order they appear in the invoice.

    Raises:
        StructuredError: If rawfiles_tpl does not exist in all indexes of ExcelInvoice, or if there are unused raw data.
    """
    # Check for the existence of rawfiles_tpl specified as an index
    # Conversely, check that all rawfiles_tpl are present in the ExcelInvoice index
    dcttpl = {str(tpl[0].parent.name): tpl for tpl in rawfiles_tpl}  # Assuming terminal folder names are unique
    dir_setglob = set(dcttpl.keys())
    dir_set_invoice = set(dfexcelinvoice["data_folder"])
    if dir_setglob == dir_set_invoice:
        # Reorder rawfiles_tpl according to the order of appearance in the invoice
        return [dcttpl[d] for d in dfexcelinvoice["data_folder"]]
    if dir_setglob - dir_set_invoice:
        emsg = f"ERROR: unused raw data: {(dir_setglob-dir_set_invoice).pop()}"
        raise StructuredError(emsg)
    if dir_set_invoice - dir_setglob:
        emsg = f"ERROR: raw data not found: {(dir_set_invoice-dir_setglob).pop()}"
        raise StructuredError(emsg)

    emsg = "ERROR: unknown error"
    raise StructuredError(emsg)  # This line should never be reached


class InvoiceFile:
    """Represents an invoice file and provides utilities to read and overwrite it.

    Attributes:
        invoice_path (Path): Path to the invoice file.
        schema_path (Path | None): Optional path to the invoice schema file used for validation.
        invoice_obj (dict): Dictionary representation of the invoice JSON file.

    Args:
        invoice_path (Path): The path to the invoice file.
        schema_path (Path | None): Optional path to the invoice schema for validation when overwriting.

    Raises:
        ValueError: If `invoice_obj` is not a dictionary.

    Example:
        # Usage
        invoice = InvoiceFile(Path("invoice.json"), schema_path=Path("invoice.schema.json"))
        invoice.invoice_obj["basic"]["dataName"] = "new_data_name"
        invoice.overwrite(Path("invoice_new.json"))
        invoice.invoice_obj["basic"]["dataName"] = "updated"
        invoice.overwrite(schema_path=Path("invoice.schema.json"))
    """

    def __init__(self, invoice_path: Path, *, schema_path: Path | None = None):
        self.invoice_path = Path(invoice_path)
        self.schema_path = Path(schema_path) if schema_path is not None else None
        self._invoice_obj = self.read()

    @property
    def invoice_obj(self) -> dict[str, Any]:
        """Gets the invoice object."""
        return self._invoice_obj

    @invoice_obj.setter
    def invoice_obj(self, value: dict[str, Any]) -> None:
        """Sets the invoice object."""
        if not isinstance(value, dict):
            emsg = "invoice_obj must be a dictionary"
            raise ValueError(emsg)
        self._invoice_obj = value

    def __getitem__(self, key: str) -> Any:
        return self._invoice_obj[key]

    def __setitem__(self, key: str, value: Any) -> None:
        self._invoice_obj[key] = value

    def __delitem__(self, key: str) -> None:
        del self._invoice_obj[key]

    def read(self, *, target_path: Path | None = None) -> dict:
        """Reads the content of the invoice file and returns it as a dictionary.

        Args:
            target_path (Optional[Path], optional): Path to the target invoice file. If not provided,
                uses the path from `self.invoice_path`. Defaults to None.

        Returns:
            dict: Dictionary representation of the invoice JSON file.
        """
        if target_path is None:
            target_path = self.invoice_path

        self.invoice_obj = readf_json(target_path)
        return self.invoice_obj

    def overwrite(
        self,
        dst_file_path: Path | None = None,
        *,
        src_obj: dict[str, Any] | Path | str | None = None,
        schema_path: Path | None = None,
    ) -> None:
        """Persist invoice data to disk.

        Args:
            dst_file_path: Destination file path. Defaults to `self.invoice_path` when omitted.
            src_obj: Source invoice data. Accepts a dict to overwrite with explicit data or a path to a JSON file.
                Defaults to the current `invoice_obj`.
            schema_path: Optional path to `invoice.schema.json` for validation. Falls back to the instance level
                `schema_path` when provided at construction time.

        Raises:
            TypeError: If `src_obj` is not a dict or path-like object.
            InvoiceSchemaValidationError: When validation fails against the provided schema.
            StructuredError: If writing the file fails.

        Note:
            When `dst_file_path` targets the instance's own `invoice_path`, the in-memory `invoice_obj` is updated with
            the sanitized data after a successful write to keep state in sync. Writing to a different destination leaves
            the instance state untouched.
        """
        destination = Path(dst_file_path) if dst_file_path is not None else self.invoice_path
        validator_schema = Path(schema_path) if schema_path is not None else self.schema_path

        if src_obj is None:
            candidate: dict[str, Any] = copy.deepcopy(self.invoice_obj)
        elif isinstance(src_obj, dict):
            candidate = copy.deepcopy(src_obj)
        elif isinstance(src_obj, (str, Path)):
            candidate = readf_json(Path(src_obj))
        else:
            emsg = "src_obj must be either a dict or a path-like object"
            raise TypeError(emsg)

        sanitized = self._sanitize_invoice_data(candidate, validator_schema)
        should_update_instance = destination == self.invoice_path
        os.makedirs(destination.parent, exist_ok=True)
        writef_json(destination, sanitized)

        if should_update_instance:
            self.invoice_obj = sanitized

    def _sanitize_invoice_data(self, candidate: dict[str, Any], schema_path: Path | None) -> dict[str, Any]:
        """Validate and normalise invoice data prior to persisting."""
        if schema_path is not None:
            from rdetoolkit.validation import InvoiceValidator

            validator = InvoiceValidator(schema_path)
            return validator.validate(obj=candidate)
        return candidate

    @classmethod
    def copy_original_invoice(cls, src_file_path: Path, dst_file_path: Path) -> None:
        """Copies the original invoice file from the source file path to the destination file path.

        Args:
            src_file_path (Path): The source file path of the original invoice file.
            dst_file_path (Path): The destination file path where the original invoice file will be copied to.

        Raises:
            StructuredError: If the source file path does not exist.

        Returns:
            None
        """
        if not os.path.exists(src_file_path):
            emsg = f"File Not Found: {src_file_path}"
            raise StructuredError(emsg)
        if src_file_path != dst_file_path:
            shutil.copy(str(src_file_path), str(dst_file_path))


class TemplateGenerator(Protocol):
    def generate(self, config: TemplateConfig) -> pd.DataFrame:
        """Generates a template based on the provided configuration.

        Args:
            config (TemplateConfig): The configuration object.

        Returns:
            pd.DataFrame: A DataFrame representing the generated template.
        """
        ...


if TYPE_CHECKING:
    if sys.version_info >= (3, 10):
        AttributeConfig = GeneralAttributeConfig | SpecificAttributeConfig
    else:
        AttributeConfig = Union[GeneralAttributeConfig, SpecificAttributeConfig]
else:
    AttributeConfig = Any


class ExcelInvoiceTemplateGenerator:
    GENERAL_PREFIX = "sample.general"
    SPECIFIC_PREFIX = "sample.specific"
    CUSTOM_PREFIX = "custom"

    def __init__(self, fixed_header: FixedHeaders):
        self.fixed_header = fixed_header

    def _version_info(self) -> pd.DataFrame:
        pd = _ensure_pandas()
        return pd.DataFrame({
            "items": ["version"],
            "values": [__version__],
        })

    def generate(self, config: TemplateConfig) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Generates a template based on the provided configuration.

        Args:
            config (TemplateConfig): The configuration object.

        Returns:
            tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
                - A DataFrame representing the generated template.
                - A DataFrame containing references for general terms.
                - A DataFrame containing references for specific terms.
                - A DataFrame containing rdetoolkit version.
        """
        pd = _ensure_pandas()
        from rdetoolkit.models.invoice_schema import InvoiceSchemaJson

        base_df = self.fixed_header.to_template_dataframe().to_pandas()
        invoice_schema_obj = readf_json(config.schema_path)
        try:
            validation_error = _ensure_validation_error()
            invoice_schema = InvoiceSchemaJson(**invoice_schema_obj)
        except validation_error as e:
            raise InvoiceSchemaValidationError(str(e)) from e
        prefixes = {
            "general": self.GENERAL_PREFIX,
            "specific": self.SPECIFIC_PREFIX,
            "custom": self.CUSTOM_PREFIX,
        }

        # Sample field
        sample_field = invoice_schema.properties.sample
        if sample_field is not None:
            _, general_term_df, specific_term_df = self._add_sample_field(base_df, config, sample_field, prefixes)

        # Custom field
        custom_field = invoice_schema.properties.custom
        if custom_field is not None:
            custom_dict = custom_field.properties.root
            for key, meta_prop in custom_dict.items():
                base_df[key] = pd.Series([None, prefixes["custom"], key, meta_prop.label.ja], index=base_df.index)

        # Select Mode: folder/file
        if config.inputfile_mode == "folder":
            first_col = base_df.columns[0]
            base_df.loc[1, first_col] = ""
            base_df.loc[2, first_col] = "data_folder"

        # Version
        version_df = self._version_info()

        return base_df, general_term_df, specific_term_df, version_df

    def _add_sample_field(self, base_df: pd.DataFrame, config: TemplateConfig, sample_field: SampleField, prefixes: dict[str, str]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        from rdetoolkit.models.invoice import (
            GeneralAttributeConfig,
            GeneralTermRegistry,
            SpecificAttributeConfig,
            SpecificTermRegistry,
        )
        from rdetoolkit.models.invoice_schema import SpecificProperty

        pd = _ensure_pandas()
        attribute_configs: list[AttributeConfig] = [
            GeneralAttributeConfig(
                type="general",
                registry=GeneralTermRegistry(str(config.general_term_path)),
                prefix=prefixes["general"],
                attributes=sample_field.properties.generalAttributes,
                requires_class_id=False,
            ),
            SpecificAttributeConfig(
                type="specific",
                registry=SpecificTermRegistry(str(config.specific_term_path)),
                prefix=prefixes["specific"],
                attributes=sample_field.properties.specificAttributes,
                requires_class_id=True,
            ),
        ]
        registered_general_terms: list[dict[str, Any]] = []
        registered_specific_terms: list[dict[str, Any]] = []
        for attr_config in attribute_configs:
            attrs = attr_config.attributes
            if not attrs or not attrs.items.root:
                if isinstance(attr_config, GeneralAttributeConfig):
                    registered_general_terms = []
                elif isinstance(attr_config, SpecificAttributeConfig):
                    registered_specific_terms = []
                continue

            for prop in attrs.items.root:
                term_id = prop.properties.term_id.const
                class_id = ""
                if isinstance(prop, SpecificProperty):
                    class_id = prop.properties.class_id.const

                try:
                    emsg = "Could not find a result corresponding to the specified term_id or class_id."
                    if isinstance(attr_config, SpecificAttributeConfig):
                        emsg = f"Could not find a result corresponding to term_id {term_id} and class_id {class_id}."
                        term = attr_config.registry.by_term_and_class_id(term_id, class_id)[0]
                        registered_specific_terms.append({
                            "sample_class_id": class_id,
                            "term_id": term_id,
                            "key_name": term["key_name"],
                        })
                    else:
                        emsg = f"Could not find a result corresponding to term_id {term_id}."
                        term = attr_config.registry.by_term_id(term_id)[0]
                        registered_general_terms.append({
                            "term_id": term_id,
                            "key_name": term["key_name"],
                        })
                except (IndexError, KeyError) as e:
                    raise StructuredError(emsg) from e

                ja_name = term["ja"]
                key_name = term["key_name"]
                name = key_name.replace(f"{attr_config.prefix}.", "")
                base_df[key_name] = [None, attr_config.prefix, name, ja_name]

        df_registered_general = pd.DataFrame(
            registered_general_terms,
            columns=["term_id", "key_name"] if not registered_general_terms else None,
        )
        df_registered_specific = pd.DataFrame(
            registered_specific_terms,
            columns=["sample_class_id", "term_id", "key_name"] if not registered_specific_terms else None,
        )

        return base_df, df_registered_general, df_registered_specific

    def save(self, dataframes: dict[str, pd.DataFrame], save_path: str) -> None:
        """Save the given DataFrame to an Excel file with specific formatting.

        Args:
            dataframes (dict[str, pd.DataFrame]): The DataFrame to be saved.
            save_path (str): The path where the Excel file will be saved.

        Note:
            The method performs the following operations:
            - Writes the DataFrame to an Excel file starting from the 5th row without headers.
            - Sets the height of the 5th row to 40.
            - Adjusts the width of all columns to 20.
            - Applies a thin border to all cells in the range from row 5 to row 40.
            - Applies a thick top border and a double bottom border to the cells in the 5th row.
        """
        pd = _ensure_pandas()
        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            for sheet_name, df in dataframes.items():
                if sheet_name != "invoice_form":
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._style_sub_sheet(writer, df, sheet_name)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    self._style_main_sheet(writer, df, sheet_name)

    def _style_main_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        border_cls, _, side_cls = _ensure_openpyxl_styles()
        get_column_letter = _ensure_openpyxl_utils()
        default_row_height: int = 40
        default_column_width: int = 20
        default_start_row: int = 4
        default_end_row: int = 41
        default_start_col: int = 1

        _ = writer.book
        worksheet = writer.sheets[sheet_name]
        worksheet.row_dimensions[4].height = default_row_height
        max_col = df.shape[1]

        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = default_column_width

        # settings cell border
        thin = side_cls(border_style="thin", color="000000")
        thick = side_cls(border_style="thick", color="000000")
        double = side_cls(border_style="double", color="000000")
        grid_border = border_cls(top=thin, left=thin, right=thin, bottom=thin)

        for row in range(default_start_row, default_end_row):
            for col in range(default_start_col, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = grid_border

        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=4, column=col)
            cell.border = border_cls(left=cell.border.left, right=cell.border.right, top=thick, bottom=double)

    def _style_sub_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        _, font_cls, _ = _ensure_openpyxl_styles()
        default_cell_style = "Normal"
        _ = writer.book
        worksheet = writer.sheets[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                cell.style = default_cell_style
                cell.font = font_cls(bold=False)


class ExcelInvoiceFile:
    """Class representing an invoice file in Excel format. Provides utilities for reading and overwriting the invoice file.

    Attributes:
        invoice_path (Path): Path to the excel invoice file (.xlsx).
        dfexcelinvoice (pd.DataFrame): Dataframe of the invoice.
        df_general (pd.DataFrame | None): Dataframe of general data (None if the sheet is absent).
        df_specific (pd.DataFrame | None): Dataframe of specific data (None if the sheet is absent).
        self.template_generator (ExcelInvoiceTemplateGenerator): Template generator for the Excelinvoice.
    """
    template_generator: ExcelInvoiceTemplateGenerator | None = None

    def __init__(self, invoice_path: Path):
        self.invoice_path = invoice_path
        self.dfexcelinvoice, self.df_general, self.df_specific = self.read()

    @classmethod
    def _get_template_generator(cls) -> ExcelInvoiceTemplateGenerator:
        if cls.template_generator is None:
            from rdetoolkit.models.invoice import FixedHeaders

            cls.template_generator = ExcelInvoiceTemplateGenerator(FixedHeaders())
        return cls.template_generator

    def read(self, *, target_path: Path | None = None) -> tuple[pd.DataFrame, pd.DataFrame | None, pd.DataFrame | None]:
        """Reads the content of the Excel invoice file and returns it as three dataframes.

        Args:
            target_path (Optional[Path], optional): Path to the excelinvoice file(.xlsx) to be read. If not provided, uses the path from `self.invoice_path`. Defaults to None.

        Returns:
            tuple[pd.DataFrame, pd.DataFrame | None, pd.DataFrame | None]:
                Three dataframes (dfexcelinvoice, df_general, df_specific). The general and specific sheets are ``None``
                when the source workbook does not define the corresponding sheet.

        Raises:
            StructuredError: If the invoice file is missing, if multiple invoice-list sheets exist, or if no
            invoice-list sheet is present.
        """
        if target_path is None:
            target_path = self.invoice_path

        if not os.path.exists(target_path):
            emsg = f"ERROR: excelinvoice not found {target_path}"
            raise StructuredError(emsg)

        pd = _ensure_pandas()
        dct_sheets = pd.read_excel(target_path, sheet_name=None, dtype=str, header=None, index_col=None)

        dfexcelinvoice, df_general, df_specific = None, None, None
        for sh_name, df in dct_sheets.items():
            if df.empty:
                continue

            sheet_type = _identify_sheet_type(sh_name, df)

            if sheet_type == "invoice":
                if dfexcelinvoice is not None:
                    emsg = "ERROR: multiple sheet in invoiceList files"
                    raise StructuredError(emsg)
                ExcelInvoiceFile.check_intermittent_empty_rows(df)
                dfexcelinvoice = _SHEET_PROCESSORS[sheet_type](df)
            elif sheet_type == "general_term":
                df_general = _SHEET_PROCESSORS[sheet_type](df)
            elif sheet_type == "specific_term":
                df_specific = _SHEET_PROCESSORS[sheet_type](df)
            # If sheet_type == "unknown", do nothing (skip)

        if dfexcelinvoice is None:
            emsg = "ERROR: no sheet in invoiceList files"
            raise StructuredError(emsg)

        return dfexcelinvoice, df_general, df_specific

    @classmethod
    def generate_template(cls, invoice_schema_path: str | Path, save_path: str | Path, file_mode: Literal["file", "folder"] = "file") -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Generates a template DataFrame based on the provided invoice schema and saves it to the specified path.

        Args:
            invoice_schema_path (str | Path): The path to the invoice schema file.
            save_path (str | Path): The path where the generated template will be saved.
            file_mode (Literal["file", "folder"], optional): The mode indicating whether the input is a file or a folder. Defaults to "file".

        Returns:
            tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
                - A DataFrame representing the generated template.
                - A DataFrame containing references for general terms.
                - A DataFrame containing references for specific terms.
        """
        from rdetoolkit.models.invoice import TemplateConfig

        config = TemplateConfig(
            schema_path=invoice_schema_path,
            general_term_path=EX_GENERALTERM,
            specific_term_path=EX_SPECIFICTERM,
            inputfile_mode=file_mode,
        )

        generator = cls._get_template_generator()
        template_df, df_general, df_specific, _df_version = generator.generate(config)
        _dataframes = {
            "invoice_form": template_df,
            "generalTerm": df_general,
            "specificTerm": df_specific,
            "_version": _df_version,
        }
        generator.save(_dataframes, str(save_path))
        return template_df, df_general, df_specific

    def save(self, save_path: str | Path, *, invoice: pd.DataFrame | None = None, sheet_name: str = "invoice_form", index: list[str] | None = None, header: list[str] | None = None) -> None:
        """Save the invoice DataFrame to an Excel file.

        Args:
            save_path (str | Path): The path where the Excel file will be saved.
            invoice (pd.DataFrame | None, optional): The DataFrame containing the invoice data. Defaults to None.
            sheet_name (str, optional): The name of the sheet in the Excel file. Defaults to "invoice_form".
            index (list[str] | None, optional): The list of index labels to use. If None, index will not be written. Defaults to None.
            header (list[str] | None, optional): The list of column headers to use. If None, header will not be written. Defaults to None.

        Returns:
            None
        """
        pd = _ensure_pandas()
        _invoice_df = invoice if invoice is not None else self.dfexcelinvoice
        try:
            if index:
                _invoice_df.index = pd.Index(index)
                _index_enabled = True
            else:
                _index_enabled = False
            if header:

                _invoice_df.columns = header
                _header_enabled = True
            else:
                _header_enabled = False
            _invoice_df.to_excel(save_path, index=_index_enabled, header=_header_enabled, sheet_name=sheet_name)
        except Exception as e:
            emsg = "Failed to save the invoice file."
            raise StructuredError(emsg) from e

    def overwrite(self, invoice_org: Path, dist_path: Path, invoice_schema_path: Path, idx: int) -> None:
        """Overwrites the content of the original invoice file based on the data from the Excel invoice and saves it as a new file.

        Args:
            invoice_org (Path): Path to the original invoice file.
            dist_path (Path): Path to where the overwritten invoice file will be saved.
            invoice_schema_path (Path): Path to the invoice schema.
            idx (int): Index of the target row in the invoice dataframe.
        """
        invoice_schema_obj = readf_json(invoice_schema_path)
        invoice_obj = readf_json(invoice_org)

        # Initialize to prevent original values from being retained when Excel invoice cells are empty.
        # Tags and related samples are not supported in this version of the Excel invoice.
        for key, value in invoice_obj.items():
            if key == "sample":
                self._initialize_sample(value)
            else:
                self._initialize_non_sample(key, value)

        for k, valstr in self.dfexcelinvoice.iloc[idx, :].dropna().items():
            self._assign_value_to_invoice(k, valstr, invoice_obj, invoice_schema_obj)

        self._ensure_sample_id_order(invoice_obj)

        writef_json(dist_path, invoice_obj, enc="utf_8")

    @staticmethod
    def check_intermittent_empty_rows(df: pd.DataFrame) -> None:
        """Function to detect if there are empty rows between data rows in the ExcelInvoice (in DataFrame format).

        If an empty row exists, an exception is raised.

        Args:
            df (pd.DataFrame): Information of Sheet 1 of ExcelInvoice.

        Raises:
            StructuredError: An exception is raised if an empty row exists.
        """
        for i, row in df.iterrows():
            if not ExcelInvoiceFile.__is_empty_row(row):
                continue
            if any(not ExcelInvoiceFile.__is_empty_row(r) for r in df.iloc[i + 1]):
                emsg = "Error! Blank lines exist between lines"
                raise StructuredError(emsg)

    @staticmethod
    def __is_empty_row(row: pd.Series) -> bool:
        pd = _ensure_pandas()
        return all(cell == "" or pd.isnull(cell) for cell in row)

    def _assign_value_to_invoice(self, key: str, value: str, invoice_obj: dict, schema_obj: dict) -> None:
        assign_funcs: dict[str, Callable[[str, str, dict[Any, Any], dict[Any, Any]], None]] = {
            "basic/": self._assign_basic,
            "sample/": self._assign_sample,
            "sample.general/": self._assign_sample_general,
            "sample.specific/": self._assign_sample_specific,
            "custom/": self._assign_custom,
        }

        for prefix, func in assign_funcs.items():
            if key.startswith(prefix):
                func(key, value, invoice_obj, schema_obj)
                break

    def _assign_basic(self, key: str, value: str, invoice_obj: dict, schema_obj: dict) -> None:
        cval = key.replace("basic/", "")
        _assign_invoice_val(invoice_obj, "basic", cval, value, schema_obj)

    def _assign_sample(self, key: str, value: str, invoice_obj: dict, schema_obj: dict) -> None:
        cval = key.replace("sample/", "")
        if cval == "names":
            _assign_invoice_val(invoice_obj, "sample", cval, [value], schema_obj)
        else:
            _assign_invoice_val(invoice_obj, "sample", cval, value, schema_obj)

    def _assign_sample_general(self, key: str, value: str, invoice_obj: dict, schema_obj: dict) -> None:
        cval = key.replace("sample.general/", "sample.general.")
        df_general = self.df_general
        if df_general is None:
            emsg = "ERROR: generalTerm sheet is required to assign general attributes."
            raise StructuredError(emsg)
        term_id = df_general[df_general["key_name"] == cval]["term_id"].values[0]
        for dictobj in invoice_obj["sample"]["generalAttributes"]:
            if dictobj.get("termId") == term_id:
                dictobj["value"] = value
                break

    def _assign_sample_specific(self, key: str, value: str, invoice_obj: dict, schema_obj: dict) -> None:
        cval = key.replace("sample.specific/", "sample.specific.")
        df_specific = self.df_specific
        if df_specific is None:
            emsg = "ERROR: specificTerm sheet is required to assign specific attributes."
            raise StructuredError(emsg)
        term_id = df_specific[df_specific["key_name"] == cval]["term_id"].values[0]
        for dictobj in invoice_obj["sample"]["specificAttributes"]:
            if dictobj.get("termId") == term_id:
                dictobj["value"] = value
                break

    def _assign_custom(self, key: str, value: str, invoice_obj: dict, schema_obj: dict) -> None:
        cval = key.replace("custom/", "")
        _assign_invoice_val(invoice_obj, "custom", cval, value, schema_obj)

    def _ensure_sample_id_order(self, invoice_obj: dict) -> None:
        sample_info_value = invoice_obj.get("sample")
        if sample_info_value is None:
            return
        if "sampleId" not in sample_info_value:
            return

        sampleid_value = invoice_obj["sample"].pop("sampleId")
        invoice_obj["sample"] = {"sampleId": sampleid_value, **invoice_obj["sample"]}

    def _initialize_sample(self, sample_obj: Any) -> None:
        for item, val in sample_obj.items():
            if item in ["sampleId", "composition", "referenceUrl", "description", "ownerId"]:
                sample_obj[item] = None
            elif item in ["generalAttributes", "specificAttributes"]:
                for attribute in val:
                    attribute["value"] = None

    def _initialize_non_sample(self, key: str, value: Any) -> None:
        if key not in ["datasetId", "sample"]:
            for item in value:
                if item not in ["dateSubmitted", "instrumentId"]:
                    value[item] = None


def backup_invoice_json_files(excel_invoice_file: Path | None, mode: str | None) -> Path:
    """Backs up invoice files and retrieves paths based on the mode specified in the input.

    For excelinvoice and rdeformat modes, it backs up invoice.json as the original file in the temp directory in MultiDataTile mode.
    For other modes, it treats the files in the invoice directory as the original files.
    After backing up, it returns the file paths for invoice_org.json and invoice.schema.json.

    Args:
        excel_invoice_file (Optional[Path]): File path for excelinvoice mode
        mode (str): mode flags

    Returns:
        tuple[Path, Path]: File paths for invoice.json and invoice.schema.json
    """
    if mode is None:
        mode = ""
    from rdetoolkit.rde2util import StorageDir

    invoice_org_filepath = StorageDir.get_specific_outputdir(False, "invoice").joinpath("invoice.json")
    if (excel_invoice_file is not None) or (mode is not None and mode.lower() in ["rdeformat", "multidatatile"]):
        invoice_org_filepath = StorageDir.get_specific_outputdir(True, "temp").joinpath("invoice_org.json")
        shutil.copy(StorageDir.get_specific_outputdir(False, "invoice").joinpath("invoice.json"), invoice_org_filepath)
    # elif mode is not None and mode.lower() in ["rdeformat", "multidatatile"]:
    #     invoice_org_filepath = StorageDir.get_specific_outputdir(True, "temp").joinpath("invoice_org.json")
    #     shutil.copy(StorageDir.get_specific_outputdir(False, "invoice").joinpath("invoice.json"), invoice_org_filepath)

    return invoice_org_filepath


def __serch_key_from_constant_variable_obj(key: str, metadata_json_obj: dict) -> dict | None:
    if key in metadata_json_obj["constant"]:
        return metadata_json_obj["constant"]
    if metadata_json_obj.get("variable"):
        _variable = metadata_json_obj["variable"]
        if len(_variable) > 0:
            return metadata_json_obj["variable"][0]
        return None
    return None


def update_description_with_features(
    rde_resource: RdeOutputResourcePath,
    dst_invoice_json: Path,
    metadata_def_json: Path,
) -> None:
    """Writes the provided features to the description field RDE.

    This function takes a dictionary of features and formats them to be written
    into the description field(to invoice.json)

    Args:
        rde_resource (RdeOutputResourcePath): Path object containing resource paths needed for RDE processing.
        dst_invoice_json (Path): Path to the invoice.json file where the features will be written.
        metadata_def_json (Path): Path to the metadata list JSON file, which may include definitions or schema information.

    Returns:
        None: The function does not return a value but writes the features to the invoice.json file in the description field.
    """
    chardet = _ensure_chardet()
    with open(dst_invoice_json, "rb") as dst_invoice:
        enc_dst_invoice_data = dst_invoice.read()
    enc = chardet.detect(enc_dst_invoice_data)["encoding"]
    with open(dst_invoice_json, encoding=enc) as f:
        invoice_obj = json.load(f)

    with open(rde_resource.invoice_schema_json, "rb") as rde_resource_invoice_schema:
        enc_rde_invoice_schema_data = rde_resource_invoice_schema.read()
    enc = chardet.detect(enc_rde_invoice_schema_data)["encoding"]
    with open(rde_resource.invoice_schema_json, encoding=enc) as f:
        invoice_schema_obj = json.load(f)

    with open(metadata_def_json, "rb") as metadata_def_json_f:
        enc_rde_invoice_schema_data = metadata_def_json_f.read()
    enc = chardet.detect(enc_rde_invoice_schema_data)["encoding"]
    with open(metadata_def_json, encoding=enc) as f:
        metadata_def_obj = json.load(f)

    with open(rde_resource.meta.joinpath("metadata.json"), encoding=enc) as f:
        metadata_json_obj = json.load(f)

    description = invoice_obj["basic"]["description"] if invoice_obj["basic"]["description"] else ""
    for key, value in metadata_def_obj.items():
        if not value.get("_feature"):
            continue

        dscheader = __serch_key_from_constant_variable_obj(key, metadata_json_obj)
        if dscheader is None:
            continue
        if dscheader.get(key) is None:
            continue

        metadata_details = value
        header_entry = dscheader[key]
        name_ja = metadata_details["name"]["ja"]
        header_value = header_entry["value"]

        if metadata_details.get("unit"):
            unit = metadata_details["unit"]
            description += f"\n{name_ja}({unit}):{header_value}"
        else:
            description += f"\n{name_ja}:{header_value}"

        if description.startswith("\n"):
            description = description[1:]

    _assign_invoice_val(invoice_obj, "basic", "description", description, invoice_schema_obj)
    writef_json(dst_invoice_json, invoice_obj)


class RuleBasedReplacer:
    """A class for changing the rules of data naming.

    This class is used to manage and apply file name mapping rules. It reads rules from a JSON format
    rule file, sets rules, and performs file name transformations and replacements based on those rules.

    Attributes:
        rules (dict[str, str]): Dictionary holding the mapping rules.
        last_apply_result (dict[str, Any]): The result of the last applied rules.

    Args:
        rule_file_path (Optional[Union[str, Path]]): Path to the rule file. If specified, rules are loaded from this path.
    """

    def __init__(self, *, rule_file_path: str | Path | None = None):
        self.rules: dict[str, str] = {}
        self.last_apply_result: dict[str, Any] = {}

        if isinstance(rule_file_path, str):
            rule_file_path = Path(rule_file_path)
        if rule_file_path and rule_file_path.exists():
            self.load_rules(rule_file_path)

    def load_rules(self, filepath: str | Path) -> None:
        """Function to read file mapping rules.

        The file containing the mapping rules must be in JSON format.

        Args:
            filepath (Union[str, Path]): The file path of the JSON file containing the mapping rules.

        Raises:
            StructuredError: An exception is raised if the file extension is not json.
        """
        if isinstance(filepath, str):
            filepath = Path(filepath)
        if filepath.suffix != ".json":
            emsg = f"Error. File format/extension is not correct: {filepath}"
            raise StructuredError(emsg)

        data = readf_json(filepath)
        self.rules = data.get("filename_mapping", {})

    def get_apply_rules_obj(
        self,
        replacements: dict[str, Any],
        source_json_obj: dict[str, Any] | None,
        *,
        mapping_rules: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        """Function to convert file mapping rules into a JSON format.

        This function takes string mappings separated by dots ('.') and converts them into a dictionary format, making it easier to handle within a target JsonObject.

        Args:
            replacements (dict[str, str]): The object containing mapping rules.
            source_json_obj (Optional[dict[str, Any]]): Objects of key and value to which you want to apply the rule
            mapping_rules (Optional[dict[str, str]], optional): Rules for mapping key and value. Defaults to None.

        Returns:
            dict[str, Any]: dictionary type data after conversion

        Example:
            # rule.json
            rule = {
                "filename_mapping": {
                    "invoice.basic.dataName": "${filename}",
                    "invoice.sample.names": ["${somedataname}"],
                }
            }
            replacer = RuleBasedReplacer('rules.json')
            replacements = {
                '${filename}': 'example.txt',
                '${somedataname}': ['some data']
            }
            result = replacer.apply_rules(replacement_rule, save_file_path, mapping_rules = rule)
            print(result)
        """
        # [TODO] Correction of type definitions in version 0.1.6
        if mapping_rules is None:
            mapping_rules = self.rules
        if source_json_obj is None:
            source_json_obj = {}

        for key, value in self.rules.items():
            keys = key.split(".")
            replace_value = replacements.get(value, "")
            current_obj: dict[str, Any] = source_json_obj
            for k in keys[:-1]:
                # search for the desired key in the dictionary from "xxx.xxx.xxx" ...
                if k not in current_obj:
                    current_obj[k] = {}
                current_obj = current_obj[k]
            current_obj[keys[-1]] = replace_value

        self.last_apply_result = source_json_obj

        return self.last_apply_result

    def set_rule(self, path: str, variable: str) -> None:
        """Sets a new rule.

        Args:
            path (str): The path to the target location for replacement.
            variable (str): The rule after replacement.

        Example:
            replacer = RuleBasedReplacer()
            replacer.set_rule('invoice.basic.dataName', 'filename')
            replacer.set_rule('invoice.sample.name', 'dataname')
            print(replacer.rules)
        """
        self.rules[path] = variable

    def write_rule(self, replacements_rule: dict[str, Any], save_file_path: str | Path) -> str:
        """Function to write file mapping rules to a target JSON file.

        Writes the set mapping rules (in JSON format) to the target file

        Args:
            replacements_rule (dict[str, str]): The object containing mapping rules.
            save_file_path (Union[str, Path]): The file path for saving.

        Raises:
            StructuredError: An exception error occurs if the extension of the save path is not .json.
            StructuredError: An exception error occurs if values cannot be written to the json.

        Returns:
            str: The result of writing to the target JSON.
        """
        contents: str = ""

        if isinstance(save_file_path, str):
            save_file_path = Path(save_file_path)

        if save_file_path.suffix != ".json":
            emsg = f"Extension error. Incorrect extension: {save_file_path}"
            raise StructuredError(emsg)

        if save_file_path.exists():
            exists_contents = readf_json(save_file_path)
            _ = self.get_apply_rules_obj(replacements_rule, exists_contents)
            data_to_write = copy.deepcopy(exists_contents)
        else:
            new_contents: dict[str, Any] = {}
            _ = self.get_apply_rules_obj(replacements_rule, new_contents)
            data_to_write = copy.deepcopy(new_contents)

        try:
            writef_json(save_file_path, data_to_write)
            contents = json.dumps({"filename_mapping": self.rules})
        except json.JSONDecodeError as json_err:
            emsg = "Error. No write was performed on the target json"
            raise StructuredError(emsg) from json_err

        return contents


class MagicVariableResolver:
    """Resolve and expand supported magic-variable expressions."""

    MIN_INVOICE_FIELD_SEGMENTS = 2
    MIN_METADATA_SEGMENTS = 2

    def __init__(
        self,
        *,
        rawfile_path: Path,
        invoice_source: dict[str, Any],
        metadata_source: dict[str, Any] | None,
    ) -> None:
        self.rawfile_path = rawfile_path
        self.invoice_source = invoice_source
        self.metadata_source = metadata_source

    def expand(self, template: str) -> str:
        """Expand all magic variables present in *template*."""
        result_parts: list[str] = []
        last_end = 0
        skip_pending = False

        for match in MAGIC_VARIABLE_PATTERN.finditer(template):
            literal = template[last_end : match.start()]
            if literal:
                literal = self._trim_redundant_underscore(literal, result_parts, skip_pending)
                result_parts.append(literal)
                skip_pending = False

            expression = match.group(1).strip()
            replacement = self._resolve_expression(expression)
            if replacement is None:
                skip_pending = True
            else:
                result_parts.append(replacement)
                skip_pending = False

            last_end = match.end()

        trailing_literal = template[last_end:]
        if trailing_literal:
            trailing_literal = self._trim_redundant_underscore(trailing_literal, result_parts, skip_pending)
            result_parts.append(trailing_literal)

        return "".join(result_parts)

    def _trim_redundant_underscore(self, literal: str, result_parts: list[str], skip_pending: bool) -> str:
        """Drop leading underscores when a skipped placeholder already supplied one."""
        if skip_pending and literal.startswith("_") and result_parts and result_parts[-1].endswith("_"):
            return literal[1:]
        return literal

    def _resolve_expression(self, expression: str) -> str | None:
        if not expression:
            emsg = "Encountered empty magic variable expression"
            raise StructuredError(emsg)

        segments = expression.split(":")
        prefix = segments[0]

        if prefix == "filename":
            return self.rawfile_path.name
        if prefix == "invoice":
            return self._resolve_invoice_expression(segments[1:], expression)
        if prefix == "metadata":
            return self._resolve_metadata_expression(segments[1:], expression)

        emsg = f"Unsupported magic variable '{expression}'"
        raise StructuredError(emsg)

    def _resolve_invoice_expression(self, segments: list[str], expression: str) -> str | None:
        if not segments:
            emsg = f"Invalid invoice magic variable '{expression}'"
            raise StructuredError(emsg)

        section = segments[0]
        invoice_section = self.invoice_source.get(section)
        if invoice_section is None:
            emsg = f"Invoice section '{section}' not found for magic variable '{expression}'"
            raise StructuredError(emsg)

        if section in {"basic", "custom"}:
            if len(segments) < self.MIN_INVOICE_FIELD_SEGMENTS:
                emsg = f"Magic variable '{expression}' requires a field name"
                raise StructuredError(emsg)
            field = segments[1]
            if not isinstance(invoice_section, dict) or field not in invoice_section:
                emsg = f"Field '{section}.{field}' is missing for magic variable '{expression}'"
                raise StructuredError(emsg)
            value = invoice_section.get(field)
            return self._normalize_scalar(value, expression)

        if section == "sample":
            return self._resolve_sample_expression(segments[1:], expression)

        emsg = f"Unsupported invoice section '{section}' in magic variable '{expression}'"
        raise StructuredError(emsg)

    def _resolve_sample_expression(self, segments: list[str], expression: str) -> str | None:
        if not segments:
            emsg = f"Magic variable '{expression}' must specify a sample field"
            raise StructuredError(emsg)

        sample_section = self.invoice_source.get("sample")
        if sample_section is None:
            emsg = f"Sample information missing in invoice for '{expression}'"
            raise StructuredError(emsg)

        field = segments[0]
        if field != "names":
            emsg = f"Unsupported sample field '{field}' in magic variable '{expression}'"
            raise StructuredError(emsg)

        names = sample_section.get("names")
        if names is None or not isinstance(names, list):
            emsg = f"'sample.names' is unavailable for magic variable '{expression}'"
            raise StructuredError(emsg)
        if len(names) == 0:
            emsg = f"Magic variable '{expression}' cannot be applied because sample.names is empty"
            raise StructuredError(emsg)

        filtered_names = [name for name in names if isinstance(name, str) and name]
        if not filtered_names:
            emsg = f"Magic variable '{expression}' cannot be applied because sample.names only contains empty strings"
            raise StructuredError(emsg)

        return "_".join(filtered_names)

    def _resolve_metadata_expression(self, segments: list[str], expression: str) -> str | None:
        if not segments:
            emsg = f"Invalid metadata magic variable '{expression}'"
            raise StructuredError(emsg)

        if segments[0] != "constant":
            emsg = f"Unsupported metadata field '{segments[0]}' in magic variable '{expression}'"
            raise StructuredError(emsg)

        if self.metadata_source is None:
            emsg = f"metadata.json is required to resolve '{expression}'"
            raise StructuredError(emsg)

        if len(segments) < self.MIN_METADATA_SEGMENTS:
            emsg = f"Magic variable '{expression}' requires a constant key"
            raise StructuredError(emsg)

        constant_key = segments[1]
        constants = self.metadata_source.get("constant", {})
        metadata_entry = constants.get(constant_key)
        if metadata_entry is None:
            emsg = f"metadata.constant['{constant_key}'] is missing for magic variable '{expression}'"
            raise StructuredError(emsg)

        return self._normalize_scalar(metadata_entry.get("value"), expression)

    def _normalize_scalar(self, value: Any, expression: str) -> str | None:
        if value is None or (isinstance(value, str) and value == ""):
            logger.warning("Magic variable '%s' resolved to an empty value and will be skipped", expression)
            return None
        if isinstance(value, (str, int, float, bool)):
            return str(value)

        emsg = f"Magic variable '{expression}' must resolve to a scalar value, got {type(value).__name__!s}"
        raise StructuredError(emsg)


def apply_default_filename_mapping_rule(replacement_rule: dict[str, Any], save_file_path: str | Path) -> dict[str, Any]:
    """Applies a default filename mapping rule based on the basename of the save file path.

    This function creates an instance of RuleBasedReplacer and applies a default mapping rule. If the basename
    of the save file path is 'invoice', it sets a specific rule for 'basic.dataName'. After setting the rule,
    it writes the mapping rule to the specified file path and returns the result of the last applied rules.

    Args:
        replacement_rule (dict[str, Any]): The replacement rules to be applied.
        save_file_path (Union[str, Path]): The file path where the replacement rules are saved.

    Returns:
        dict[str, Any]: The result of the last applied replacement rules.

    The function assumes the existence of certain structures in the replacement rules and file paths, and it
    specifically checks for a basename of 'invoice' to apply a predefined rule.
    """
    if isinstance(save_file_path, str):
        basename = os.path.splitext(os.path.basename(save_file_path))[0]
    elif isinstance(save_file_path, Path):
        basename = save_file_path.stem

    replacer = RuleBasedReplacer()
    if basename == "invoice":
        replacer.set_rule("basic.dataName", "${filename}")
    replacer.write_rule(replacement_rule, save_file_path)

    return replacer.last_apply_result


def _load_metadata(dataset_paths: RdeDatasetPaths | None) -> dict[str, Any] | None:
    """Load metadata.json when dataset paths are available.

    Args:
        dataset_paths: Dataset paths that may include a metadata directory.

    Returns:
        Parsed metadata contents when the file exists; otherwise None.
    """
    if dataset_paths is None:
        return None

    metadata_path = dataset_paths.meta.joinpath("metadata.json")
    if not metadata_path.exists():
        return None

    return readf_json(metadata_path)


def apply_magic_variable(
    invoice_path: str | Path,
    rawfile_path: str | Path,
    *,
    save_filepath: str | Path | None = None,
    dataset_paths: RdeDatasetPaths | None = None,
) -> dict[str, Any]:
    """Expand supported magic variables inside ``invoice.json``.

    Magic variables can reference the input filename, invoice metadata sourced
    from ``invoice_org`` and constants defined in ``metadata.json``.  Only
    ``basic.dataName`` currently supports substitution.

    Args:
        invoice_path: Target invoice file to update.
        rawfile_path: Raw input file supplying ``${filename}``.
        save_filepath: Destination path for the updated invoice. Defaults to ``invoice_path``.
        dataset_paths: Dataset paths used to locate ``invoice_org`` and metadata files.

    Returns:
        dict[str, Any]: Updated invoice contents when substitutions occur. An empty dict is returned when
        no magic variables are present.

    Raises:
        StructuredError: If required fields or metadata are missing for a referenced magic variable.
    """
    invoice_path = Path(invoice_path)
    rawfile_path = Path(rawfile_path)
    destination_path = Path(save_filepath) if save_filepath is not None else invoice_path

    invoice_contents = readf_json(invoice_path)
    basic_section = invoice_contents.get("basic")
    if basic_section is None:
        emsg = "invoice.json is missing the 'basic' section required for magic variable processing"
        raise StructuredError(emsg)

    data_name_template = basic_section.get("dataName")
    if not isinstance(data_name_template, str) or MAGIC_VARIABLE_PATTERN.search(data_name_template) is None:
        # No magic variables to apply.
        return {}

    invoice_source_path = dataset_paths.invoice_org if dataset_paths is not None else invoice_path
    invoice_source = readf_json(invoice_source_path)

    metadata_contents = _load_metadata(dataset_paths)

    resolver = MagicVariableResolver(
        rawfile_path=rawfile_path,
        invoice_source=invoice_source,
        metadata_source=metadata_contents,
    )
    resolved_name = resolver.expand(data_name_template)
    if resolved_name == "":
        emsg = "Magic variable expansion produced an empty dataName"
        raise StructuredError(emsg)

    basic_section["dataName"] = resolved_name
    writef_json(destination_path, invoice_contents)
    return invoice_contents


class SmartTableFile:
    """Handles SmartTable files (Excel/CSV/TSV) for invoice generation.

    This class reads table files containing metadata and maps them to invoice structure
    according to SmartTable format specifications. The first row (display names) is
    skipped, and the second row is used as the mapping key headers.
    """

    def __init__(self, smarttable_path: Path):
        """Initialize SmartTableFile with the path to the table file.

        Args:
            smarttable_path: Path to the SmartTable file (.xlsx, .csv, .tsv)

        Raises:
            StructuredError: If file format is not supported or file doesn't exist.
        """
        self.smarttable_path = smarttable_path
        self._validate_file()
        self._data: pd.DataFrame | None = None

    def _validate_file(self) -> None:
        """Validate the SmartTable file format and existence."""
        if not self.smarttable_path.exists():
            error_msg = f"SmartTable file not found: {self.smarttable_path}"
            raise StructuredError(error_msg)

        supported_extensions = [".xlsx", ".csv", ".tsv"]
        if self.smarttable_path.suffix.lower() not in supported_extensions:
            error_msg = f"Unsupported file format: {self.smarttable_path.suffix}. Supported formats: {supported_extensions}"
            raise StructuredError(error_msg)

        if not self.smarttable_path.name.startswith("smarttable_"):
            error_msg = f"Invalid naming convention: {self.smarttable_path.name}. File must start with 'smarttable_'"
            raise StructuredError(error_msg)

    def read_table(self) -> pd.DataFrame:
        """Read the SmartTable file and return as DataFrame.

        The first row (display names) is skipped, and the second row is used
        as column headers (mapping keys).

        Returns:
            DataFrame containing the table data with mapping key headers.

        Raises:
            StructuredError: If file reading fails or format is invalid.
        """
        if self._data is not None:
            return self._data

        pd = _ensure_pandas()
        try:
            if self.smarttable_path.suffix.lower() == ".xlsx":
                # Read Excel file, skip first row (display names), use second row as header
                self._data = pd.read_excel(self.smarttable_path, sheet_name=0, dtype=str, skiprows=[0], header=0)
            elif self.smarttable_path.suffix.lower() == ".csv":
                # Read CSV file, skip first row (display names), use second row as header
                self._data = pd.read_csv(self.smarttable_path, dtype=str, skiprows=[0], header=0)
            elif self.smarttable_path.suffix.lower() == ".tsv":
                # Read TSV file, skip first row (display names), use second row as header
                self._data = pd.read_csv(self.smarttable_path, sep="\t", dtype=str, skiprows=[0], header=0)

            if self._data is None:
                error_msg = (
                    "Unsupported SmartTable file extension: "
                    f"{self.smarttable_path.suffix}"
                )
                raise StructuredError(error_msg)

            mapping_prefixes = ["basic/", "custom/", "sample/", "meta/", "inputdata"]
            has_mapping_keys = any(
                any(col.startswith(prefix) for prefix in mapping_prefixes)
                for col in self._data.columns
            )
            if not has_mapping_keys:
                error_msg = "SmartTable file must have mapping keys with prefixes: basic/, custom/, sample/, meta/, inputdata"
                raise StructuredError(error_msg)

            return self._data

        except Exception as e:
            if isinstance(e, StructuredError):
                raise
            error_msg = f"Failed to read SmartTable file {self.smarttable_path}: {str(e)}"
            raise StructuredError(error_msg) from e

    def generate_row_csvs_with_file_mapping(
        self,
        output_dir: Path,
        extracted_files: list[Path] | None = None,
    ) -> list[tuple[Path, tuple[Path, ...]]]:
        """Generate individual CSV files for each row with file mapping.

        This method creates a CSV file for each data row and maps inputdata<N> columns
        to actual file paths from extracted zip files.

        Args:
            output_dir: Directory to save individual CSV files.
            extracted_files: List of extracted files from zip (optional).

        Returns:
            List of tuples where each tuple contains:
            - Path to the generated CSV file
            - Tuple of related file paths based on inputdata<N> columns

        Raises:
            StructuredError: If CSV generation or file mapping fails.
        """
        data = self.read_table()
        csv_file_mappings = []

        pd = _ensure_pandas()
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            inputdata_columns = [col for col in data.columns if col.startswith("inputdata")]

            for idx, row in data.iterrows():
                # New naming convention: smarttable_<original_filename>_XXXX.csv
                csv_filename = f"f{self.smarttable_path.stem}_{idx:04d}.csv"
                csv_path = output_dir / csv_filename

                single_row_df = pd.DataFrame([row], columns=data.columns)
                single_row_df.to_csv(csv_path, index=False)

                # Map inputdata<N> values to actual file paths
                related_files = []
                for col in inputdata_columns:
                    file_relative_path = row[col]
                    if pd.isna(file_relative_path) or file_relative_path == "":
                        continue

                    # Find matching file in extracted files
                    if extracted_files:
                        matching_file = self._find_file_by_relative_path(
                            file_relative_path, extracted_files,
                        )
                        if matching_file:
                            related_files.append(matching_file)

                csv_file_mappings.append((csv_path, tuple(related_files)))
            return csv_file_mappings

        except Exception as e:
            error_msg = f"Failed to generate CSV files with file mapping: {str(e)}"
            raise StructuredError(error_msg) from e

    def _find_file_by_relative_path(self, relative_path: str, extracted_files: list[Path]) -> Path | None:
        """Find extracted file by relative path.

        Args:
            relative_path: Relative path from inputdata<N> column.
            extracted_files: List of extracted files.

        Returns:
            Path to the matching file or None if not found.
        """
        # Normalize the relative path (remove leading/trailing slashes)
        normalized_path = relative_path.strip("/\\")

        for file_path in extracted_files:
            file_str = str(file_path)

            if file_str.replace("\\", "/").endswith(normalized_path.replace("\\", "/")):
                return file_path

            # Also check by exact relative path match
            # Extract relative part after temp directory
            path_parts = file_path.parts
            min_parts = 2  # at least temp/file
            if len(path_parts) >= min_parts:
                relative_part = "/".join(path_parts[-len(Path(normalized_path).parts):])
                if relative_part == normalized_path.replace("\\", "/"):
                    return file_path
        return None
