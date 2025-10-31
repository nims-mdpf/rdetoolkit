"""Test system files filtering in input checkers.

Test that system files like .DS_Store are properly filtered out
and don't cause validation errors.
"""

from pathlib import Path
import shutil
import zipfile
import pytest

from rdetoolkit.impl.input_controller import (
    ExcelInvoiceChecker,
    InvoiceChecker,
    MultiFileChecker,
    RDEFormatChecker,
)


@pytest.fixture
def setup_test_dir():
    """Setup and cleanup test directory."""
    test_dir = Path("data")
    test_dir.mkdir(exist_ok=True)
    yield test_dir
    if test_dir.exists():
        shutil.rmtree(test_dir)


@pytest.fixture
def inputdata_with_ds_store(setup_test_dir):
    """Create input directory with .DS_Store file."""
    input_dir = setup_test_dir / "inputdata"
    input_dir.mkdir(exist_ok=True)
    
    # Create .DS_Store file (system file that should be filtered)
    ds_store = input_dir / ".DS_Store"
    ds_store.write_text("fake DS_Store content")
    
    return input_dir


@pytest.fixture
def excel_invoice_file(inputdata_with_ds_store):
    """Create a minimal Excel invoice file."""
    excel_file = inputdata_with_ds_store / "sample_excel_invoice.xlsx"
    
    # Create a minimal Excel file using openpyxl
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "basic/dataName"
        ws["A2"] = "test_data"
        wb.save(str(excel_file))
    except ImportError:
        # If openpyxl is not available, create a simple file as placeholder
        excel_file.touch()
    
    return excel_file


@pytest.fixture
def zip_file_with_data(inputdata_with_ds_store):
    """Create a zip file with test data."""
    zip_path = inputdata_with_ds_store / "data.zip"
    
    # Create a temporary file to zip
    temp_file = Path("test_data.txt")
    temp_file.write_text("test content")
    
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(temp_file)
    
    temp_file.unlink()
    return zip_path


class TestExcelInvoiceCheckerWithSystemFiles:
    """Test ExcelInvoiceChecker handles system files correctly."""
    
    def test_ds_store_filtered_out_before_validation(self, inputdata_with_ds_store, excel_invoice_file):
        """Test that .DS_Store is filtered out and doesn't cause 'other files' error."""
        unpacked_dir = Path("data/temp")
        unpacked_dir.mkdir(parents=True, exist_ok=True)
        
        checker = ExcelInvoiceChecker(unpacked_dir)
        
        # Get the grouped files - .DS_Store should be filtered before grouping
        input_files = list(inputdata_with_ds_store.glob("*"))
        
        # First verify .DS_Store exists in the directory
        ds_store = inputdata_with_ds_store / ".DS_Store"
        assert ds_store.exists()
        assert ds_store in input_files
        
        # Now filter using SystemFilesCleaner as the parse method does
        from rdetoolkit.impl.compressed_controller import SystemFilesCleaner
        cleaner = SystemFilesCleaner()
        filtered_files = [f for f in input_files if not cleaner.is_excluded(f)]
        
        # .DS_Store should be filtered out
        assert ds_store not in filtered_files
        
        # Now check that grouping works without .DS_Store
        zipfiles, excel_files, other_files = checker._get_group_by_files(filtered_files)
        
        # Should have the Excel file and no other files (DS_Store was filtered)
        assert len(excel_files) == 1
        assert excel_invoice_file in excel_files
        assert len(other_files) == 0  # This would have been 1 if .DS_Store wasn't filtered
    
    def test_ds_store_with_zip_filtered(self, inputdata_with_ds_store, excel_invoice_file, zip_file_with_data):
        """Test that .DS_Store is filtered when both zip and Excel files are present."""
        unpacked_dir = Path("data/temp")
        unpacked_dir.mkdir(parents=True, exist_ok=True)
        
        checker = ExcelInvoiceChecker(unpacked_dir)
        
        # Get the grouped files
        input_files = list(inputdata_with_ds_store.glob("*"))
        
        # Filter using SystemFilesCleaner
        from rdetoolkit.impl.compressed_controller import SystemFilesCleaner
        cleaner = SystemFilesCleaner()
        filtered_files = [f for f in input_files if not cleaner.is_excluded(f)]
        
        # Verify .DS_Store is filtered
        ds_store = inputdata_with_ds_store / ".DS_Store"
        assert ds_store not in filtered_files
        
        # Now check grouping
        zipfiles, excel_files, other_files = checker._get_group_by_files(filtered_files)
        
        # Should have the Excel file, zip file, and no other files
        assert len(excel_files) == 1
        assert len(zipfiles) == 1
        assert len(other_files) == 0  # .DS_Store filtered out, so this is empty


class TestInvoiceCheckerWithSystemFiles:
    """Test InvoiceChecker handles system files correctly."""
    
    def test_parse_with_ds_store(self, inputdata_with_ds_store):
        """Test that .DS_Store is filtered out in invoice mode."""
        # Create a regular file
        regular_file = inputdata_with_ds_store / "test_file.txt"
        regular_file.write_text("test content")
        
        unpacked_dir = Path("data/temp")
        unpacked_dir.mkdir(parents=True, exist_ok=True)
        
        checker = InvoiceChecker(unpacked_dir)
        rawfiles, _ = checker.parse(inputdata_with_ds_store)
        
        assert len(rawfiles) == 1
        # Only the regular file should be in rawfiles, not .DS_Store
        assert regular_file in rawfiles[0]
        assert inputdata_with_ds_store / ".DS_Store" not in rawfiles[0]


class TestMultiFileCheckerWithSystemFiles:
    """Test MultiFileChecker handles system files correctly."""
    
    def test_parse_with_ds_store(self, inputdata_with_ds_store):
        """Test that .DS_Store is filtered out in multifile mode."""
        # Create regular files
        file1 = inputdata_with_ds_store / "file1.txt"
        file2 = inputdata_with_ds_store / "file2.txt"
        file1.write_text("content1")
        file2.write_text("content2")
        
        unpacked_dir = Path("data/temp")
        unpacked_dir.mkdir(parents=True, exist_ok=True)
        
        checker = MultiFileChecker(unpacked_dir)
        rawfiles, _ = checker.parse(inputdata_with_ds_store)
        
        # Only the two regular files should be included
        assert len(rawfiles) == 2
        file_paths = [rf[0] for rf in rawfiles]
        assert file1 in file_paths
        assert file2 in file_paths
        assert inputdata_with_ds_store / ".DS_Store" not in file_paths


class TestExcelExtensionMatching:
    """Test that Excel file extensions are correctly matched."""
    
    def test_xls_extension_recognized(self, setup_test_dir):
        """Test that .xls files are recognized as Excel invoice files."""
        input_dir = setup_test_dir / "inputdata"
        input_dir.mkdir(exist_ok=True)
        
        # Create .xls file
        xls_file = input_dir / "test_excel_invoice.xls"
        xls_file.touch()
        
        checker = InvoiceChecker(Path("data/temp"))
        zipfiles, excel_files, other_files = checker._get_group_by_files([xls_file])
        
        assert len(excel_files) == 1
        assert xls_file in excel_files
        assert len(other_files) == 0
    
    def test_xlsx_extension_recognized(self, setup_test_dir):
        """Test that .xlsx files are recognized as Excel invoice files."""
        input_dir = setup_test_dir / "inputdata"
        input_dir.mkdir(exist_ok=True)
        
        # Create .xlsx file
        xlsx_file = input_dir / "test_excel_invoice.xlsx"
        xlsx_file.touch()
        
        checker = InvoiceChecker(Path("data/temp"))
        zipfiles, excel_files, other_files = checker._get_group_by_files([xlsx_file])
        
        assert len(excel_files) == 1
        assert xlsx_file in excel_files
        assert len(other_files) == 0
    
    def test_uppercase_xlsx_extension_recognized(self, setup_test_dir):
        """Test that .XLSX files are recognized (case insensitive)."""
        input_dir = setup_test_dir / "inputdata"
        input_dir.mkdir(exist_ok=True)
        
        # Create .XLSX file with uppercase extension
        xlsx_file = input_dir / "test_excel_invoice.XLSX"
        xlsx_file.touch()
        
        checker = InvoiceChecker(Path("data/temp"))
        zipfiles, excel_files, other_files = checker._get_group_by_files([xlsx_file])
        
        assert len(excel_files) == 1
        assert xlsx_file in excel_files
        assert len(other_files) == 0
